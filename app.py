import os
import sqlite3
import hashlib
from flask import Flask, request, redirect, url_for, session, send_file, render_template_string, flash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from datetime import datetime
import cloudinary
import cloudinary.uploader
from fpdf import FPDF

# --- CONFIG ---
app = Flask(__name__)
app.secret_key = "supersecretkey"
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Cloudinary (ambil dari environment Render)
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET")
)

# --- DB INIT ---
def init_db():
    with sqlite3.connect("rekap.db") as conn:
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        kode_toko TEXT UNIQUE,
                        password TEXT,
                        role TEXT DEFAULT 'user'
                    )""")
        c.execute("""CREATE TABLE IF NOT EXISTS pengiriman (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        nrb TEXT,
                        tgl_pengiriman TEXT,
                        no_mobil TEXT,
                        driver TEXT,
                        bukti_url TEXT
                    )""")
init_db()

# --- UTILS ---
def hash_password(pw): return hashlib.sha256(pw.encode()).hexdigest()
def get_user_id(): return session.get("user_id")
def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*a, **kw):
        if not session.get("user_id"): return redirect(url_for("login"))
        return f(*a, **kw)
    return wrapper

# --- BASE HTML ---
def base_template(content):
    return render_template_string(f"""
    <!doctype html>
    <html lang="id">
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>Rekap Pengiriman</title>
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
      <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap" rel="stylesheet">
      <style>
        body {{
          font-family:'Poppins',sans-serif;
          background:#f4f6f9;
          transition: background 0.3s, color 0.3s;
        }}
        .card {{
          border-radius:1rem;
          box-shadow:0 4px 12px rgba(0,0,0,.1);
          opacity:0;
          transform: translateY(15px);
          animation: fadeInUp 0.6s ease forwards;
        }}
        .table-responsive {{
          opacity:0;
          animation: fadeIn 0.8s ease forwards;
          animation-delay: .3s;
        }}
        @keyframes fadeInUp {{
          from {{ opacity:0; transform: translateY(15px); }}
          to   {{ opacity:1; transform: translateY(0); }}
        }}
        @keyframes fadeIn {{
          from {{ opacity:0; }}
          to   {{ opacity:1; }}
        }}
        /* Toast */
        .toast {{
          opacity:0;
          transition: opacity 0.5s ease-in-out, transform 0.4s ease-in-out;
          transform: translateX(20px);
        }}
        .toast.show {{
          opacity:1;
          transform: translateX(0);
        }}
        /* Dark Mode */
        body.dark-mode {{
          background: #1e1e2f;
          color: #f5f5f5;
        }}
        body.dark-mode .card {{
          background: #2c2c3c;
          color: #f5f5f5;
          box-shadow: 0 4px 12px rgba(255,255,255,.05);
        }}
        body.dark-mode .table {{
          color: #f5f5f5;
        }}
        body.dark-mode .btn-outline-dark {{
          color: #f5f5f5;
          border-color: #f5f5f5;
        }}
        body.dark-mode .btn-outline-dark:hover {{
          background: #f5f5f5;
          color: #2c2c3c;
        }}
        /* Row animation highlight */
        @keyframes rowFadeIn {{
          0%   {{ background-color: #ffeaa7; opacity: 0; transform: translateY(-5px); }}
          50%  {{ background-color: #fdcb6e; opacity: 1; }}
          100% {{ background-color: transparent; opacity: 1; transform: translateY(0); }}
        }}
        .new-row {{
          animation: rowFadeIn 1.2s ease forwards;
        }}
      </style>
    </head>
    <body>
      <div class="container py-4">
        <div class="d-flex justify-content-end mb-3">
          <button id="darkToggle" class="btn btn-outline-dark">🌙 Dark Mode</button>
        </div>
        {content}
      </div>

      <!-- Toast Container -->
      <div class="toast-container position-fixed top-0 end-0 p-3" style="z-index: 1100;">
        <div id="mainToast" class="toast align-items-center text-bg-success border-0" role="alert">
          <div class="d-flex">
            <div class="toast-body" id="toastMessage">✅ Data berhasil disimpan!</div>
            <button type="button" class="btn-close btn-close-white me-2 m-auto" data-bs-dismiss="toast"></button>
          </div>
        </div>
      </div>

      <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
      <script>
        // Toast helper
        function showToast(message, type="success") {{
          const toastEl = document.getElementById("mainToast");
          const toastBody = document.getElementById("toastMessage");
          toastBody.textContent = message;
          toastEl.className = "toast align-items-center border-0 text-bg-" + 
                              (type === "error" ? "danger" : "success");
          const toast = new bootstrap.Toast(toastEl, {{ delay: 3000, autohide: true }});
          toast.show();
        }}

        // Dark mode toggle
        const toggleBtn = document.getElementById("darkToggle");
        if (localStorage.getItem("theme") === "dark") {{
          document.body.classList.add("dark-mode");
          toggleBtn.textContent = "☀️ Light Mode";
        }}
        toggleBtn.addEventListener("click", () => {{
          document.body.classList.toggle("dark-mode");
          if (document.body.classList.contains("dark-mode")) {{
            toggleBtn.textContent = "☀️ Light Mode";
            localStorage.setItem("theme","dark");
          }} else {{
            toggleBtn.textContent = "🌙 Dark Mode";
            localStorage.setItem("theme","light");
          }}
        }});

        // Flash messages to toast
        document.addEventListener("DOMContentLoaded", () => {{
          {% with messages = get_flashed_messages(with_categories=true) %}
            {% for category, message in messages %}
              showToast("{{ message }}", "{{ category }}");
            {% endfor %}
          {% endwith %}
        }});
      </script>
    </body>
    </html>
    """)

# --- AUTH ---
@app.route("/register", methods=["GET","POST"])
def register():
    if request.method=="POST":
        kode = request.form["kode_toko"]
        pw = hash_password(request.form["password"])
        try:
            with sqlite3.connect("rekap.db") as conn:
                conn.execute("INSERT INTO users (kode_toko,password) VALUES (?,?)",(kode,pw))
            flash("Registrasi berhasil, silakan login ✅","success")
            return redirect(url_for("login"))
        except:
            flash("❌ Kode toko sudah terdaftar!","error")
            return redirect(url_for("register"))
    return base_template("""
    <h3>Registrasi</h3>
    <form method="POST" class="card p-4">
      <input name="kode_toko" placeholder="Kode Toko" class="form-control mb-2" required>
      <input type="password" name="password" placeholder="Password" class="form-control mb-2" required>
      <button class="btn btn-primary w-100">Daftar</button>
      <a href="/login">Login</a>
    </form>
    """)

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method=="POST":
        kode = request.form["kode_toko"]; pw = hash_password(request.form["password"])
        with sqlite3.connect("rekap.db") as conn:
            c=conn.cursor()
            c.execute("SELECT id,role FROM users WHERE kode_toko=? AND password=?",(kode,pw))
            u=c.fetchone()
        if u:
            session.update({"user_id":u[0],"kode_toko":kode,"role":u[1]})
            flash("Login berhasil ✅","success")
            return redirect(url_for("index"))
        flash("❌ Login gagal, periksa kode/password","error")
        return redirect(url_for("login"))
    return base_template("""
    <h3>Login</h3>
    <form method="POST" class="card p-4">
      <input name="kode_toko" placeholder="Kode Toko" class="form-control mb-2" required>
      <input type="password" name="password" placeholder="Password" class="form-control mb-2" required>
      <button class="btn btn-success w-100">Login</button>
      <a href="/register">Daftar</a>
    </form>
    """)

@app.route("/logout")
def logout(): 
    session.clear()
    flash("Berhasil logout 👋","success")
    return redirect(url_for("login"))

# --- DASHBOARD ---
@app.route("/", methods=["GET","POST"])
@login_required
def index():
    user_id=get_user_id(); search=request.args.get("search_nrb","").strip()
    if request.method=="POST":
        nrb=request.form["nrb"]; tgl=request.form["tgl_pengiriman"]
        no=request.form["no_mobil"]; drv=request.form["driver"]
        bukti=request.files["bukti"]; url=""
        if bukti and bukti.filename:
            fname=secure_filename(bukti.filename); path=os.path.join(UPLOAD_FOLDER,fname)
            bukti.save(path); up=cloudinary.uploader.upload(path); url=up.get("secure_url","")
        with sqlite3.connect("rekap.db") as conn:
            conn.execute("INSERT INTO pengiriman (user_id,nrb,tgl_pengiriman,no_mobil,driver,bukti_url) VALUES (?,?,?,?,?,?)",
                         (user_id,nrb,tgl,no,drv,url))
        flash("Data berhasil disimpan ✅","success")
        return redirect(url_for("index"))

    with sqlite3.connect("rekap.db") as conn:
        c=conn.cursor()
        if session["role"]=="admin":
            if search:
                c.execute("""SELECT u.kode_toko,p.nrb,p.tgl_pengiriman,p.no_mobil,p.driver,p.bukti_url
                             FROM pengiriman p JOIN users u ON p.user_id=u.id WHERE p.nrb LIKE ? ORDER BY p.id DESC""",(f"%{search}%",))
            else:
                c.execute("""SELECT u.kode_toko,p.nrb,p.tgl_pengiriman,p.no_mobil,p.driver,p.bukti_url
                             FROM pengiriman p JOIN users u ON p.user_id=u.id ORDER BY p.id DESC""")
        else:
            if search:
                c.execute("SELECT nrb,tgl_pengiriman,no_mobil,driver,bukti_url FROM pengiriman WHERE user_id=? AND nrb LIKE ? ORDER BY id DESC",(user_id,f"%{search}%"))
            else:
                c.execute("SELECT nrb,tgl_pengiriman,no_mobil,driver,bukti_url FROM pengiriman WHERE user_id=? ORDER BY id DESC",(user_id,))
        data=c.fetchall()

    rows=""
    for i,r in enumerate(data):
        cls="new-row" if i==0 else ""
        if session["role"]=="admin":
            rows+=f"<tr class='{cls}'><td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td><td>{r[4]}</td><td><a href='{r[5]}' target='_blank'>Bukti</a></td></tr>"
        else:
            rows+=f"<tr class='{cls}'><td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td><td><a href='{r[4]}' target='_blank'>Bukti</a></td></tr>"

    return base_template(f"""
    <h2 class="text-center">📦 Rekap Pengiriman<br><small>{session['kode_toko']} ({session['role']})</small></h2>
    <div class="d-flex flex-column flex-md-row justify-content-between mb-3 gap-2">
      <a href="/logout" class="btn btn-danger">Logout</a>
      <div class="d-flex flex-column flex-sm-row gap-2">
        <a href="/export?search_nrb={search}" class="btn btn-success">📤 Excel</a>
        <a href="/export_pdf?search_nrb={search}" class="btn btn-danger">📑 PDF</a>
      </div>
    </div>
    <div class="card p-3 mb-3">
      <form method="POST" enctype="multipart/form-data" class="row g-2">
        <div class="col-12 col-md-3"><input name="nrb" placeholder="NRB" class="form-control" required></div>
        <div class="col-12 col-md-3"><input type="date" name="tgl_pengiriman" class="form-control" required></div>
        <div class="col-12 col-md-3"><input name="no_mobil" placeholder="Nomor Mobil" class="form-control" required></div>
        <div class="col-12 col-md-3"><input name="driver" placeholder="Nama Driver" class="form-control" required></div>
        <div class="col-12"><input type="file" name="bukti" class="form-control"></div>
        <div class="col-12"><button class="btn btn-primary w-100">💾 Simpan</button></div>
      </form>
    </div>
    <form method="GET" class="row g-2 mb-3">
      <div class="col-12 col-md-4"><input name="search_nrb" value="{search}" class="form-control" placeholder="Cari NRB"></div>
      <div class="col-12 col-md-2"><button class="btn btn-secondary w-100">Cari</button></div>
    </form>
    <div class="card p-3 table-responsive">
      <table class="table table-bordered table-hover">
        <thead><tr>{("<th>Kode Toko</th>" if session['role']=="admin" else "")}<th>NRB</th><th>Tanggal</th><th>No Mobil</th><th>Driver</th><th>Bukti</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
    """)

# --- EXPORT EXCEL ---
@app.route("/export")
@login_required
def export():
    user_id=get_user_id(); search=request.args.get("search_nrb","").strip()
    with sqlite3.connect("rekap.db") as conn:
        c=conn.cursor()
        if session["role"]=="admin":
            q="SELECT u.kode_toko,p.nrb,p.tgl_pengiriman,p.no_mobil,p.driver,p.bukti_url FROM pengiriman p JOIN users u ON p.user_id=u.id"
            if search: q+=" WHERE p.nrb LIKE ? ORDER BY p.id DESC"; c.execute(q,(f"%{search}%",))
            else: q+=" ORDER BY p.id DESC"; c.execute(q)
        else:
            q="SELECT nrb,tgl_pengiriman,no_mobil,driver,bukti_url FROM pengiriman WHERE user_id=?"
            if search: q+=" AND nrb LIKE ? ORDER BY id DESC"; c.execute(q,(user_id,f"%{search}%",))
            else: q+=" ORDER BY id DESC"; c.execute(q,(user_id,))
        data=c.fetchall()
    wb=Workbook(); ws=wb.active; ws.title="Rekap"
    ws.append(["Kode Toko","NRB","Tanggal","No Mobil","Driver","Bukti"] if session["role"]=="admin" else ["NRB","Tanggal","No Mobil","Driver","Bukti"])
    for row in data: ws.append(row)
    fn=f"rekap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"; path=f"/tmp/{fn}"; wb.save(path)
    return send_file(path,as_attachment=True,download_name=fn)

# --- EXPORT PDF ---
@app.route("/export_pdf")
@login_required
def export_pdf():
    user_id=get_user_id(); search=request.args.get("search_nrb","").strip()
    with sqlite3.connect("rekap.db") as conn:
        c=conn.cursor()
        if session["role"]=="admin":
            q="SELECT u.kode_toko,p.nrb,p.tgl_pengiriman,p.no_mobil,p.driver FROM pengiriman p JOIN users u ON p.user_id=u.id"
            if search: q+=" WHERE p.nrb LIKE ? ORDER BY p.id DESC"; c.execute(q,(f"%{search}%",))
            else: q+=" ORDER BY p.id DESC"; c.execute(q)
        else:
            q="SELECT nrb,tgl_pengiriman,no_mobil,driver FROM pengiriman WHERE user_id=?"
            if search: q+=" AND nrb LIKE ? ORDER BY id DESC"; c.execute(q,(user_id,f"%{search}%",))
            else: q+=" ORDER BY id DESC"; c.execute(q,(user_id,))
        data=c.fetchall()
    pdf=FPDF(); pdf.add_page(); pdf.set_font("Arial","B",14); pdf.cell(0,10,"Rekap Pengiriman",0,1,"C"); pdf.set_font("Arial","",10)
    colw=[30,30,30,40,40]; header=["Kode Toko","NRB","Tanggal","No Mobil","Driver"] if session["role"]=="admin" else ["NRB","Tanggal","No Mobil","Driver"]
    for i,h in enumerate(header): pdf.cell(colw[i],8,h,1,0,"C"); pdf.ln()
    for r in data:
        for i,c in enumerate(r): pdf.cell(colw[i],8,str(c),1)
        pdf.ln()
    fn=f"rekap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"; path=f"/tmp/{fn}"; pdf.output(path)
    return send_file(path,as_attachment=True,download_name=fn)

if __name__=="__main__": app.run(debug=True)
